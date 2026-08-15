"""设备管理 API：设备 CRUD + 凭据登记 + 触发采集 + 采集状态。

安全约定：
- 凭据只用于登记（加密存储），API 只返回是否配置与算法摘要；
- 写操作 require_write；凭据相关操作 require_admin；
- 采集接口触发只读采集，绝不改设备配置。
"""
import asyncio

from fastapi import APIRouter, Depends, HTTPException, Query, status
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.deps import get_current_user, require_operator_admin, require_write, require_admin
from app.models.device import Device, DeviceConfigSnapshot, DeviceCredential, SnmpInterfaceMetric
from app.models.user import User
from app.schemas.common import Response, PageData
from app.schemas.device import (
    ConfigChangeEventOut,
    ConfigDiffOut,
    ConfigDiffRow,
    DeviceCollectRequest,
    DeviceCollectResponse,
    DeviceCollectStatus,
    DeviceConfigCollectOut,
    DeviceConfigSnapshotOut,
    DeviceConfigTextOut,
    DeviceCredentialIn,
    DeviceCredentialOut,
    DeviceIn,
    DeviceLldpCollectOut,
    DeviceListOut,
    DeviceListResponse,
    DeviceOut,
    DeviceResponse,
    InterfaceTrendOut,
    LldpNeighborOut,
    SnmpInterfaceOut,
    SnmpInterfaceResponse,
)
from app.services import credential_manager
from app.services.device_collector import collect_device, collect_devices

router = APIRouter(prefix="/api/devices", tags=["devices"])


def _cred_out(cred: DeviceCredential) -> DeviceCredentialOut:
    return DeviceCredentialOut(
        id=cred.id,
        name=cred.name,
        protocol=cred.protocol,
        username=cred.username,
        auth_algorithm=cred.auth_algorithm,
        priv_algorithm=cred.priv_algorithm,
        has_secret=bool(
            cred.auth_key_encrypted or cred.priv_key_encrypted or cred.ssh_key_encrypted
        ),
        external_secret_ref=cred.external_secret_ref or "",
        created_at=cred.created_at,
    )


def _device_out(db: Session, device: Device) -> DeviceOut:
    data = DeviceOut.model_validate(device)
    data.has_snmp = bool(device.snmp_config_id)
    data.has_ssh = bool(device.ssh_config_id)
    return data


# ---- 凭据 ----

@router.post("/credentials", response_model=Response[DeviceCredentialOut],
             status_code=status.HTTP_201_CREATED)
def create_credential(
    payload: DeviceCredentialIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> Response[DeviceCredentialOut]:
    cred = DeviceCredential(
        name=payload.name,
        protocol=payload.protocol,
        username=payload.username,
        auth_algorithm=payload.auth_algorithm,
        priv_algorithm=payload.priv_algorithm,
        external_secret_ref=payload.external_secret_ref,
    )
    # 加密存储（未配置 NETCHECK_SECRET_KEY 时存空标记，由 collector 报错）
    if payload.auth_key:
        cred.auth_key_encrypted = credential_manager.encrypt_secret(payload.auth_key)
    if payload.priv_key:
        cred.priv_key_encrypted = credential_manager.encrypt_secret(payload.priv_key)
    if payload.ssh_key:
        cred.ssh_key_encrypted = credential_manager.encrypt_secret(payload.ssh_key)
    db.add(cred)
    db.commit()
    db.refresh(cred)
    return Response(data=_cred_out(cred))


@router.get("/credentials", response_model=Response[PageData[DeviceCredentialOut]])
def list_credentials(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response[PageData[DeviceCredentialOut]]:
    query = db.query(DeviceCredential)
    total = query.count()
    items = (
        query.order_by(DeviceCredential.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return Response(data=PageData(
        total=total, page=page, page_size=page_size,
        items=[_cred_out(c) for c in items],
    ))


@router.delete("/credentials/{credential_id}", response_model=Response)
def delete_credential(
    credential_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
) -> Response:
    cred = db.query(DeviceCredential).filter(DeviceCredential.id == credential_id).first()
    if cred is None:
        raise HTTPException(status_code=404, detail="凭据不存在")
    db.delete(cred)
    db.commit()
    return Response()


# ---- 设备 ----

@router.post("", response_model=DeviceResponse, status_code=status.HTTP_201_CREATED)
def create_device(
    payload: DeviceIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> DeviceResponse:
    device = Device(
        asset_id=payload.asset_id,
        name=payload.name,
        management_ip=payload.management_ip,
        vendor_platform=payload.vendor_platform,
        snmp_config_id=payload.snmp_config_id,
        ssh_config_id=payload.ssh_config_id,
    )
    db.add(device)
    db.commit()
    db.refresh(device)
    return Response(data=_device_out(db, device))


@router.get("", response_model=DeviceListResponse)
def list_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    vendor: str | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DeviceListResponse:
    query = db.query(Device)
    if vendor:
        query = query.filter(Device.vendor_platform == vendor)
    total = query.count()
    items = (
        query.order_by(Device.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return Response(data=PageData(
        total=total, page=page, page_size=page_size,
        items=[DeviceListOut.model_validate(d) for d in items],
    ))


@router.get("/{device_id}", response_model=DeviceResponse)
def get_device(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DeviceResponse:
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    return Response(data=_device_out(db, device))


@router.get("/{device_id}/interfaces", response_model=SnmpInterfaceResponse)
def list_device_interfaces(
    device_id: int,
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> SnmpInterfaceResponse:
    if db.query(Device).filter(Device.id == device_id).first() is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    metrics = (
        db.query(SnmpInterfaceMetric)
        .filter(SnmpInterfaceMetric.device_id == device_id)
        .order_by(SnmpInterfaceMetric.interface_index.asc(),
                  SnmpInterfaceMetric.collected_at.desc())
        .limit(limit)
        .all()
    )
    return Response(data=[SnmpInterfaceOut.model_validate(m) for m in metrics])


@router.put("/{device_id}", response_model=DeviceResponse)
def update_device(
    device_id: int,
    payload: DeviceIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> DeviceResponse:
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    device.name = payload.name
    device.management_ip = payload.management_ip
    device.vendor_platform = payload.vendor_platform
    device.snmp_config_id = payload.snmp_config_id
    device.ssh_config_id = payload.ssh_config_id
    if payload.asset_id is not None:
        device.asset_id = payload.asset_id
    db.commit()
    db.refresh(device)
    return Response(data=_device_out(db, device))


@router.delete("/{device_id}", response_model=Response)
def delete_device(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> Response:
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    # 清理关联指标与配置快照（N2.1 P1：级联/清理策略）
    db.query(SnmpInterfaceMetric).filter(
        SnmpInterfaceMetric.device_id == device_id
    ).delete()
    db.query(DeviceConfigSnapshot).filter(
        DeviceConfigSnapshot.device_id == device_id
    ).delete()
    db.delete(device)
    db.commit()
    return Response()


@router.post("/collect", response_model=DeviceCollectResponse)
def trigger_collect(
    payload: DeviceCollectRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> DeviceCollectResponse:
    """触发指定设备采集（最多 8 台，同步执行）。"""
    results = collect_devices(db, payload.device_ids)
    first = results[0] if results else {"status": "error", "error": "无设备"}
    return Response(data=DeviceCollectStatus.model_validate(first))


@router.post("/{device_id}/collect", response_model=DeviceCollectResponse)
def trigger_collect_one(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> DeviceCollectResponse:
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    result = collect_device(db, device)
    return Response(data=DeviceCollectStatus.model_validate(result))

# ---- N2 配置备份与差异 ----

@router.get(
    "/{device_id}/configs",
    response_model=Response[PageData[DeviceConfigSnapshotOut]],
)
def list_config_snapshots(
    device_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> Response[PageData[DeviceConfigSnapshotOut]]:
    """列出设备的配置快照（升序，最新在后）。"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    query = (
        db.query(DeviceConfigSnapshot)
        .filter(DeviceConfigSnapshot.device_id == device_id)
        .order_by(DeviceConfigSnapshot.collected_at.asc())
    )
    total = query.count()
    items = (
        query.offset((page - 1) * page_size).limit(page_size).all()
    )
    return Response(
        data=PageData(
            total=total,
            page=page,
            page_size=page_size,
            items=[DeviceConfigSnapshotOut.model_validate(i) for i in items],
        )
    )


@router.get("/{device_id}/configs/latest", response_model=Response[DeviceConfigTextOut])
def get_latest_config(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator_admin),
) -> Response[DeviceConfigTextOut]:
    """获取最新配置快照（脱敏文本）。无快照返回 404。仅 operator/admin 可读。"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    snap = (
        db.query(DeviceConfigSnapshot)
        .filter(DeviceConfigSnapshot.device_id == device_id)
        .order_by(DeviceConfigSnapshot.collected_at.desc(),
                  DeviceConfigSnapshot.id.desc())
        .first()
    )
    if snap is None:
        raise HTTPException(status_code=404, detail="尚无配置快照")
    return Response(data=DeviceConfigTextOut(
        id=snap.id,
        device_id=snap.device_id,
        vendor_platform=snap.vendor_platform,
        config_full_hash=snap.config_full_hash,
        config_text_redacted=snap.config_text_redacted,
        source=snap.source,
        changed=snap.changed,
        truncated=snap.truncated,
        collected_at=snap.collected_at,
    ))


@router.post("/{device_id}/configs/collect", response_model=Response[DeviceConfigCollectOut])
def trigger_config_collect(
    device_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_write),
) -> Response[DeviceConfigCollectOut]:
    """触发单台设备的配置备份采集（同步，HttpOnly 凭据加密）。"""
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    from app.services.config_backup import collect_config_snapshot

    result = asyncio.run(collect_config_snapshot(db, device))
    # 审计日志
    from app.models.audit import OperationLog

    detail = f"config_backup status={result.get('status')}"
    if result.get("hash"):
        detail += f" sha256={result['hash'][:8]}"
    if result.get("changed"):
        detail += " changed=True"
    if result.get("truncated"):
        detail += " truncated=True"
    audit = OperationLog(
        username=current_user.username,
        action="device_config_backup",
        target_type="device",
        target_id=device_id,
        detail=detail,
    )
    db.add(audit)
    db.commit()
    return Response(data=DeviceConfigCollectOut(
        device_id=device_id,
        status=result.get("status", "error"),
        snapshot_id=result.get("snapshot_id"),
        changed=result.get("changed"),
        hash=result.get("hash"),
        command=result.get("command"),
        truncated=result.get("truncated"),
        error=result.get("error"),
    ))


@router.get("/{device_id}/configs/diff", response_model=Response[ConfigDiffOut])
def diff_configs_endpoint(
    device_id: int,
    from_snapshot_id: int | None = Query(None),
    to_snapshot_id: int | None = Query(None),
    context_lines: int = Query(3, ge=0, le=50, description="上下文行数（每侧）"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_operator_admin),
) -> Response[ConfigDiffOut]:
    """对比两个配置快照。默认对比最新两份；可显式 from/to。

    diff 行数上限：settings.config_diff_max_rows（默认 2000），超出标记 capped=True。
    from 必须早于 to（按时间），否则 400。仅 operator/admin 可读。
    """
    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    q = (
        db.query(DeviceConfigSnapshot)
        .filter(DeviceConfigSnapshot.device_id == device_id)
        .order_by(DeviceConfigSnapshot.collected_at.desc(),
                  DeviceConfigSnapshot.id.desc())
    )
    snaps = q.limit(50).all()
    if not snaps:
        raise HTTPException(status_code=404, detail="尚无配置快照")

    from app.services.config_backup import diff_configs

    if from_snapshot_id and to_snapshot_id:
        snaps_map = {s.id: s for s in snaps}
        older = snaps_map.get(from_snapshot_id)
        newer = snaps_map.get(to_snapshot_id)
        if older is None or newer is None:
            raise HTTPException(status_code=404, detail="快照不存在")
    elif len(snaps) >= 2:
        newer = snaps[0]          # 最新
        older = snaps[1]          # 上一份
    else:
        raise HTTPException(status_code=400, detail="快照不足两份，无法对比")

    # 时间窗校验：from 必须早于 to（按 (collected_at, id) 排序）
    def _order_key(s):
        return (s.collected_at, s.id)

    if _order_key(older) >= _order_key(newer):
        raise HTTPException(status_code=400, detail="from 必须早于 to（按时间）")

    rows = diff_configs(older.config_text_redacted, newer.config_text_redacted,
                        context_lines=context_lines)
    from app.services.config_backup import format_diff_text
    from app.core.config import settings

    max_rows = settings.config_diff_max_rows
    capped = len(rows) > max_rows
    if capped:
        rows = rows[:max_rows]
    text = format_diff_text(rows)
    # 行数限制
    text_lines = text.splitlines()
    if len(text_lines) > max_rows:
        text_lines = text_lines[:max_rows]
        text = "\n".join(text_lines)
        capped = True

    return Response(data=ConfigDiffOut(
        device_id=device_id,
        from_snapshot_id=older.id,
        to_snapshot_id=newer.id,
        from_collected_at=older.collected_at,
        to_collected_at=newer.collected_at,
        changed=older.config_full_hash != newer.config_full_hash,
        rows=[ConfigDiffRow(**r) for r in rows],
        text=text,
        capped=capped,
    ))


# ---- N4 网络可观测闭环 ----

@router.get("/{device_id}/interfaces/trend")
def interface_trend(
    device_id: int,
    start: str = Query(..., description="ISO 起始时间"),
    end: str = Query(..., description="ISO 结束时间"),
    interval: int = Query(60, ge=1, le=86400, description="聚合桶秒数"),
    interface_index: int | None = Query(None, description="指定接口，缺省全部"),
    db: Session = Depends(get_db),
    _: User = Depends(require_operator_admin),
):
    """N4 接口指标趋势（append-only 样本表聚合，缺样本返回 null 不补 0）。"""
    from app.services.trend_service import (
        MAX_TREND_SPAN_SECONDS,
        parse_iso,
        query_interface_trend,
    )

    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    from app.services.trend_service import DEFAULT_MAX_POINTS

    start_dt = parse_iso(start)
    end_dt = parse_iso(end)
    if start_dt is None or end_dt is None:
        raise HTTPException(status_code=400, detail="时间格式错误（示例 2026-08-15T00:00:00）")
    from datetime import timedelta

    if end_dt - start_dt > timedelta(seconds=MAX_TREND_SPAN_SECONDS):
        raise HTTPException(status_code=400, detail=f"查询跨度超过上限 {MAX_TREND_SPAN_SECONDS//3600} 小时")
    data = query_interface_trend(db, device_id, interface_index, start_dt, end_dt,
                                 interval, DEFAULT_MAX_POINTS)
    return Response(data=data)


@router.get("/{device_id}/lldp", response_model=Response[list[LldpNeighborOut]])
def lldp_neighbors(
    device_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_operator_admin),
):
    """N4 LLDP 邻居观测列表（最近活动）。"""
    from app.models.lldp import LldpObservation

    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    rows = (
        db.query(LldpObservation)
        .filter(LldpObservation.device_id == device_id)
        .order_by(LldpObservation.local_port_index.asc(), LldpObservation.last_seen.desc())
        .all()
    )
    return Response(data=[LldpNeighborOut.model_validate(r) for r in rows])


@router.post("/{device_id}/lldp/collect", response_model=Response[DeviceLldpCollectOut])
def lldp_collect(
    device_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(require_operator_admin),
):
    """N4 立即采集一次 LLDP 邻居（只读 SNMP WALK）。"""
    from app.services.lldp_collector import collect_lldp_neighbors

    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    if not device.snmp_config_id:
        raise HTTPException(status_code=400, detail="设备未配置 SNMPv3 凭据")
    result = collect_lldp_neighbors(db, device)
    return Response(data=DeviceLldpCollectOut(
        device_id=device_id,
        status=result.get("status", "error"),
        neighbors=result.get("neighbors", 0),
        stored=result.get("stored", 0),
        error=result.get("error"),
    ))


@router.get("/{device_id}/configs/events")
def config_change_events(
    device_id: int,
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
    _: User = Depends(require_operator_admin),
):
    """N4 配置变化事件列表（独立于巡检 run 的审计事实）。"""
    from app.models.device import ConfigChangeEvent

    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    rows = (
        db.query(ConfigChangeEvent)
        .filter(ConfigChangeEvent.device_id == device_id)
        .order_by(ConfigChangeEvent.triggered_at.desc(), ConfigChangeEvent.id.desc())
        .limit(limit)
        .all()
    )
    return Response(data=[ConfigChangeEventOut.model_validate(r) for r in rows])


@router.get("/{device_id}/configs/diff/export")
def config_diff_export(
    device_id: int,
    fmt: str = Query("text", pattern="^(text|excel)$"),
    from_id: int | None = Query(None, description="缺省用倒数第二份快照"),
    to_id: int | None = Query(None, description="缺省用最新快照"),
    db: Session = Depends(get_db),
    _: User = Depends(require_operator_admin),
):
    """N4 配置差异导出（text / excel）。纯读操作，数据来自已持久化快照。"""
    from fastapi.responses import PlainTextResponse, StreamingResponse
    from io import BytesIO

    from app.core.config import settings
    from app.services.config_backup import diff_configs, format_diff_text
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    device = db.query(Device).filter(Device.id == device_id).first()
    if device is None:
        raise HTTPException(status_code=404, detail="设备不存在")
    snaps = (
        db.query(DeviceConfigSnapshot)
        .filter(DeviceConfigSnapshot.device_id == device_id)
        .order_by(DeviceConfigSnapshot.collected_at.desc(),
                  DeviceConfigSnapshot.id.desc())
        .all()
    )
    if len(snaps) < 2:
        raise HTTPException(status_code=400, detail="快照不足两份，无法导出差异")
    if to_id is not None:
        newer = next((s for s in snaps if s.id == to_id), None)
        if newer is None:
            raise HTTPException(status_code=404, detail="to 快照不存在")
    else:
        newer = snaps[0]
    if from_id is not None:
        older = next((s for s in snaps if s.id == from_id), None)
        if older is None:
            raise HTTPException(status_code=404, detail="from 快照不存在")
    else:
        older = snaps[1]
    if (older.collected_at, older.id) >= (newer.collected_at, newer.id):
        raise HTTPException(status_code=400, detail="from 必须早于 to")

    rows = diff_configs(older.config_text_redacted, newer.config_text_redacted)
    max_rows = settings.config_diff_max_rows
    capped = len(rows) > max_rows
    if capped:
        rows = rows[:max_rows]
    text = format_diff_text(rows)

    import re as _re

    def _sanitize_filename(name: str) -> str:
        safe = _re.sub(r"[^A-Za-z0-9._-]+", "_", name or "device")
        return safe[:48] or "device"

    base = f"{_sanitize_filename(device.name)}_config_diff_{older.id}_{newer.id}"
    if fmt == "text":
        payload = (text + ("\n[!] 差异已截断（超过上限）" if capped else "")).encode("utf-8")
        return PlainTextResponse(
            content=payload,
            media_type="text/plain; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{base}.txt"'},
        )

    # Excel：行类型着色；防公式注入（= + - @ 开头转为文本）
    wb = Workbook()
    ws = wb.active
    ws.title = "config-diff"
    headers = ["kind", "old_line", "new_line", "text"]
    ws.append(headers)
    for col_idx, _ in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)
    fill_map = {
        "add": PatternFill(fill_type="solid", start_color="C6EFCE", end_color="C6EFCE"),
        "del": PatternFill(fill_type="solid", start_color="FFC7CE", end_color="FFC7CE"),
        "context": None,
        "skip": PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9"),
    }
    for r in rows:
        kind = r.get("kind", "context")
        text_val = r.get("text", "")
        if text_val.startswith(("=", "+", "-", "@")):
            text_val = "'" + text_val  # 防 Excel 公式注入
        ws.append([kind, r.get("old_line_no"), r.get("new_line_no"), text_val])
        fill = fill_map.get(kind)
        if fill is not None:
            row_idx = ws.max_row
            for col_idx in range(1, 5):
                ws.cell(row=row_idx, column=col_idx).fill = fill
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 100
    if capped:
        ws.append(["skip", None, None, "[!] 差异已截断（超过上限）"])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{base}.xlsx"'},
    )
