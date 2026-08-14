from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.services.checkers import CheckResult
from app.services.checkers import CHECKERS as CHECKERS_IMPL

from helpers import wait_run


def _headers(token: str) -> dict[str, str]:
    return {"Authorization": f"Bearer {token}"}


def _create_run_with_abnormal(client: TestClient, auth_token: str, monkeypatch):
    headers = _headers(auth_token)
    task = client.post(
        "/api/tasks",
        headers=headers,
        json={"name": "闭环巡检", "check_types": ["http"], "asset_ids": [1, 2]},
    ).json()["data"]

    class HttpChecker:
        def check(self, asset):
            if asset.id == 1:
                return [CheckResult("failed", f"http://{asset.ip}", 120, error_message="HTTP 500")]
            return [CheckResult("warning", f"http://{asset.ip}", 3000, "HTTP 200，响应缓慢")]

    monkeypatch.setitem(CHECKERS_IMPL, "http", HttpChecker())
    run = client.post(f"/api/tasks/{task['id']}/run", headers=headers).json()["data"]
    wait_run(client, headers, run["id"])
    return headers, task, run


def test_dashboard_requires_token(client: TestClient):
    assert client.get("/api/dashboard/summary").status_code == 401


def test_dashboard_summary_fault_types_and_recent_abnormal(client: TestClient, auth_token: str, monkeypatch):
    headers, task, run = _create_run_with_abnormal(client, auth_token, monkeypatch)

    summary = client.get("/api/dashboard/summary", headers=headers)
    assert summary.status_code == 200
    data = summary.json()["data"]
    assert data["asset_total"] == 12
    assert data["task_total"] == 1
    assert data["run_total"] == 1
    assert data["today_abnormal_results"] == 2
    assert data["diagnosis_total"] == 2

    fault_types = client.get("/api/dashboard/fault-types?days=7", headers=headers)
    assert fault_types.status_code == 200
    names = {item["name"] for item in fault_types.json()["data"]}
    assert {"Web应用内部错误", "网络拥塞或服务性能下降"}.issubset(names)

    recent = client.get("/api/dashboard/recent-abnormal?limit=1", headers=headers)
    assert recent.status_code == 200
    recent_items = recent.json()["data"]
    assert len(recent_items) == 1
    assert recent_items[0]["run_id"] == run["id"]
    assert recent_items[0]["task_id"] == task["id"]
    assert recent_items[0]["task_name"] == "闭环巡检"
    assert recent_items[0]["status"] in {"failed", "warning"}


def test_global_results_list_and_filters(client: TestClient, auth_token: str, monkeypatch):
    headers, task, run = _create_run_with_abnormal(client, auth_token, monkeypatch)

    all_results = client.get("/api/results", headers=headers)
    assert all_results.status_code == 200
    data = all_results.json()["data"]
    assert data["total"] == 2
    assert {"task_id", "task_name", "asset_name"}.issubset(data["items"][0])

    filtered = client.get(
        "/api/results",
        headers=headers,
        params={"run_id": run["id"], "task_id": task["id"], "asset_id": 1, "check_type": "http", "status": "failed"},
    )
    assert filtered.status_code == 200
    filtered_data = filtered.json()["data"]
    assert filtered_data["total"] == 1
    item = filtered_data["items"][0]
    assert item["asset_id"] == 1
    assert item["task_name"] == "闭环巡检"
    assert item["status"] == "failed"


def test_reports_generate_list_download_delete_and_excel_content(client: TestClient, auth_token: str, monkeypatch):
    headers, task, run = _create_run_with_abnormal(client, auth_token, monkeypatch)

    generated = client.post("/api/reports/generate", headers=headers, json={"report_type": "run", "run_id": run["id"]})
    assert generated.status_code == 201
    report = generated.json()["data"]
    assert report["report_type"] == "run"
    assert report["run_id"] == run["id"]
    assert report["task_id"] == task["id"]
    assert report["file_size"] > 0

    reports = client.get("/api/reports", headers=headers)
    assert reports.status_code == 200
    assert reports.json()["data"]["total"] == 1

    download = client.get(f"/api/reports/{report['id']}/download", headers=headers)
    assert download.status_code == 200
    assert download.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    workbook = load_workbook(BytesIO(download.content))
    sheet = workbook.active
    values = [cell for row in sheet.iter_rows(values_only=True) for cell in row if cell is not None]
    assert "巡检概况" in values
    assert "异常资产" in values
    assert "故障类型" in values
    assert "处理建议" in values
    assert "Web应用内部错误" in values
    assert "网络拥塞或服务性能下降" in values

    deleted = client.delete(f"/api/reports/{report['id']}", headers=headers)
    assert deleted.status_code == 200
    assert deleted.json()["message"] == "报告已删除"
    assert client.get(f"/api/reports/{report['id']}/download", headers=headers).status_code == 404
